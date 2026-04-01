"""Export service for Excel and PDF reports."""
import sqlite3
from datetime import date
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from terminal_prime.database.invoice_repo import InvoiceRepo
from terminal_prime.database.payment_repo import PaymentRepo


class ExportService:
    def __init__(self, conn: sqlite3.Connection):
        self.conn = conn
        self.invoice_repo = InvoiceRepo(conn)
        self.payment_repo = PaymentRepo(conn)

    def export_aging_excel(self, filepath: str, today: Optional[date] = None) -> None:
        """Export aging report to Excel with styled headers and auto column widths."""
        today = today or date.today()
        wb = Workbook()
        ws = wb.active
        ws.title = "Balance Agee"

        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F538D", end_color="1F538D", fill_type="solid")

        headers = ["N.Facture", "Client", "Montant", "Paye", "Restant", "Jours Retard", "Statut"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data rows - get all invoices (large limit)
        invoices = self.invoice_repo.get_all(limit=10000)
        for row_idx, inv in enumerate(invoices, 2):
            days_overdue = (today - inv.due_date).days
            ws.cell(row=row_idx, column=1, value=inv.number)
            ws.cell(row=row_idx, column=2, value=inv.client_name or "")
            ws.cell(row=row_idx, column=3, value=inv.amount)
            ws.cell(row=row_idx, column=4, value=inv.total_paid)
            ws.cell(row=row_idx, column=5, value=inv.remaining)
            ws.cell(row=row_idx, column=6, value=max(days_overdue, 0))
            ws.cell(row=row_idx, column=7, value=inv.display_status())

        # Auto column widths
        for col_idx in range(1, len(headers) + 1):
            max_width = len(headers[col_idx - 1])
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_width = max(max_width, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = max_width + 4

        wb.save(filepath)

    def export_invoice_pdf(self, invoice_id: int, filepath: str) -> None:
        """Export a single invoice to PDF using reportlab."""
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet

        inv = self.invoice_repo.get_by_id(invoice_id)
        if inv is None:
            raise ValueError(f"Invoice {invoice_id} not found")

        doc = SimpleDocTemplate(filepath, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        # Header
        elements.append(Paragraph("Terminal Prime - Facture", styles["Title"]))
        elements.append(Spacer(1, 20))

        # Invoice details
        detail_data = [
            ["N. Facture", inv.number],
            ["Client", inv.client_name or "N/A"],
            ["Affilie", inv.affiliate_name or "N/A"],
            ["Date", inv.date.strftime("%d/%m/%Y")],
            ["Echeance", inv.due_date.strftime("%d/%m/%Y")],
            ["Montant", f"{inv.amount:,} FCFA".replace(",", " ")],
            ["Paye", f"{inv.total_paid:,} FCFA".replace(",", " ")],
            ["Restant", f"{inv.remaining:,} FCFA".replace(",", " ")],
            ["Statut", inv.display_status()],
        ]
        detail_table = Table(detail_data, colWidths=[150, 300])
        detail_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), colors.Color(0.12, 0.33, 0.55)),
            ("TEXTCOLOR", (0, 0), (0, -1), colors.white),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(detail_table)

        # Payments table
        payments = self.payment_repo.get_by_invoice(invoice_id)
        if payments:
            elements.append(Spacer(1, 30))
            elements.append(Paragraph("Paiements", styles["Heading2"]))
            elements.append(Spacer(1, 10))

            pay_data = [["Reference", "Date", "Montant", "Mode"]]
            for pay in payments:
                pay_data.append([
                    pay.reference,
                    pay.date.strftime("%d/%m/%Y"),
                    f"{pay.amount:,} FCFA".replace(",", " "),
                    pay.mode.value,
                ])
            pay_table = Table(pay_data, colWidths=[120, 100, 120, 100])
            pay_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.12, 0.33, 0.55)),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]))
            elements.append(pay_table)

        doc.build(elements)
