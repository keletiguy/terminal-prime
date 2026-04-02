"""Import service for Mediciel Excel files.

Uses raw XML extraction from .xlsx to bypass openpyxl style parsing bugs
(Mediciel exports contain 'biltinId' typo that crashes openpyxl).
"""
import sqlite3
import zipfile
import xml.etree.ElementTree as ET
from datetime import date, datetime, timedelta
from typing import Any, Dict, List

from terminal_prime.database.client_repo import ClientRepo
from terminal_prime.database.affiliate_repo import AffiliateRepo
from terminal_prime.database.invoice_repo import InvoiceRepo
from terminal_prime.database.payment_repo import PaymentRepo


# Excel serial date epoch (1899-12-30 for the 1900 date system)
_EXCEL_EPOCH = date(1899, 12, 30)

# Mediciel column indices (0-based, after header row)
_COL_DATE = 2        # "Date facture"
_COL_CLIENT = 3      # "Client principal"
_COL_AFFILIATE = 4   # "Affilié"
_COL_NUMBER = 5      # "N° Facture"
_COL_AMOUNT = 6      # "Montant Total"
_COL_REGLEMENT = 7   # "Règlement"

_NS = {"s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


def _read_xlsx_raw(file_path: str) -> List[List[str]]:
    """Read xlsx by parsing XML directly, bypassing openpyxl style bugs."""
    with zipfile.ZipFile(file_path, "r") as z:
        # Read shared strings
        strings = []
        if "xl/sharedStrings.xml" in z.namelist():
            tree = ET.parse(z.open("xl/sharedStrings.xml"))
            for si in tree.findall(".//s:si", _NS):
                parts = []
                for t in si.iter(f"{{{_NS['s']}}}t"):
                    if t.text:
                        parts.append(t.text)
                strings.append("".join(parts))

        # Read first sheet
        tree = ET.parse(z.open("xl/worksheets/sheet1.xml"))
        xml_rows = tree.findall(".//s:sheetData/s:row", _NS)

        rows = []
        for xml_row in xml_rows:
            cells = xml_row.findall("s:c", _NS)
            values = []
            for cell in cells:
                cell_type = cell.get("t", "")
                val_elem = cell.find("s:v", _NS)
                val = val_elem.text if val_elem is not None else ""
                if cell_type == "s" and val:
                    idx = int(val)
                    val = strings[idx] if idx < len(strings) else val
                values.append(val)
            rows.append(values)

    return rows


def _parse_date(value: Any) -> date:
    """Parse a date value from various formats."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        return _EXCEL_EPOCH + timedelta(days=int(value))
    if isinstance(value, str):
        value = value.strip()
        if not value:
            raise ValueError("Empty date string")
        try:
            return _EXCEL_EPOCH + timedelta(days=int(float(value)))
        except ValueError:
            return date.fromisoformat(value)
    raise ValueError(f"Cannot parse date from {value!r}")


class ImportService:
    def __init__(self, conn: sqlite3.Connection):
        self.conn = conn
        self.client_repo = ClientRepo(conn)
        self.affiliate_repo = AffiliateRepo(conn)
        self.invoice_repo = InvoiceRepo(conn)
        self.payment_repo = PaymentRepo(conn)

    def import_file(self, file_path: str) -> Dict[str, int]:
        """Import invoices from a Mediciel Excel file.

        Returns dict with keys: imported, duplicates, clients_created,
        affiliates_created, errors.
        """
        stats = {
            "imported": 0,
            "duplicates": 0,
            "updated": 0,
            "clients_created": 0,
            "affiliates_created": 0,
            "errors": 0,
        }

        # Track which clients/affiliates existed before this import
        existing_clients: Dict[str, int] = {}
        existing_affiliates: Dict[str, int] = {}

        # Try openpyxl first, fall back to raw XML if it crashes (Mediciel bug)
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            rows = [row for row in ws.iter_rows(min_row=2, values_only=True)]
            wb.close()
        except TypeError:
            # Mediciel exports have 'biltinId' typo that crashes openpyxl
            all_rows = _read_xlsx_raw(file_path)
            rows = all_rows[1:]  # skip header row

        for row in rows:
            try:
                # Skip empty rows
                if row is None or len(row) < _COL_AMOUNT + 1:
                    continue

                invoice_number = row[_COL_NUMBER]
                if not invoice_number:
                    continue

                invoice_number = str(invoice_number).strip()
                if not invoice_number:
                    continue

                # Parse amount
                amount_val = row[_COL_AMOUNT]
                if not amount_val:
                    stats["errors"] += 1
                    continue
                amount = int(float(amount_val))

                # Parse reglement (column 7)
                reglement = 0
                if len(row) > _COL_REGLEMENT and row[_COL_REGLEMENT]:
                    try:
                        reglement = int(float(row[_COL_REGLEMENT]))
                    except (ValueError, TypeError):
                        reglement = 0

                # Check if invoice already exists
                existing = self.conn.execute(
                    "SELECT id, client_id FROM invoices WHERE number = ?",
                    (invoice_number,),
                ).fetchone()

                if existing:
                    # Invoice exists — check if reglement changed
                    inv_id = existing["id"]
                    client_id = existing["client_id"]
                    current_paid = self.conn.execute(
                        "SELECT COALESCE(SUM(amount), 0) FROM payments WHERE invoice_id = ? AND reference LIKE 'MEDICIEL-%'",
                        (inv_id,),
                    ).fetchone()[0]

                    if reglement != current_paid:
                        # Delete old Mediciel payments and create new one
                        self.conn.execute(
                            "DELETE FROM payments WHERE invoice_id = ? AND reference LIKE 'MEDICIEL-%'",
                            (inv_id,),
                        )
                        if reglement > 0:
                            self.conn.execute(
                                "INSERT INTO payments (invoice_id, client_id, date, amount, mode, reference) VALUES (?, ?, ?, ?, ?, ?)",
                                (inv_id, client_id, date.today().isoformat(), reglement, "VIREMENT", f"MEDICIEL-{invoice_number}"),
                            )
                        self.conn.commit()
                        self.invoice_repo.update_status_from_payments(inv_id)
                        stats["updated"] += 1
                    else:
                        stats["duplicates"] += 1
                    continue

                # Parse date
                date_val = row[_COL_DATE]
                if date_val is None:
                    stats["errors"] += 1
                    continue
                inv_date = _parse_date(date_val)

                # Parse client
                client_name = row[_COL_CLIENT]
                if not client_name:
                    stats["errors"] += 1
                    continue
                client_name = str(client_name).strip()

                # Parse affiliate
                affiliate_name = row[_COL_AFFILIATE]
                if not affiliate_name:
                    stats["errors"] += 1
                    continue
                affiliate_name = str(affiliate_name).strip()

                # Get or create client
                if client_name in existing_clients:
                    client_id = existing_clients[client_name]
                else:
                    existing_row = self.conn.execute(
                        "SELECT id FROM clients WHERE name = ?",
                        (client_name,),
                    ).fetchone()
                    if existing_row:
                        client_id = existing_row["id"]
                    else:
                        client = self.client_repo.create(client_name)
                        client_id = client.id
                        stats["clients_created"] += 1
                    existing_clients[client_name] = client_id

                # Get or create affiliate
                aff_key = (affiliate_name, client_id)
                if aff_key in existing_affiliates:
                    affiliate_id = existing_affiliates[aff_key]
                else:
                    existing_row = self.conn.execute(
                        "SELECT id FROM affiliates WHERE name = ? AND client_id = ?",
                        (affiliate_name, client_id),
                    ).fetchone()
                    if existing_row:
                        affiliate_id = existing_row["id"]
                    else:
                        affiliate = self.affiliate_repo.create(affiliate_name, client_id)
                        affiliate_id = affiliate.id
                        stats["affiliates_created"] += 1
                    existing_affiliates[aff_key] = affiliate_id

                # Calculate due date
                due_date = inv_date + timedelta(days=30)

                # Create invoice
                self.invoice_repo.create(
                    number=invoice_number,
                    client_id=client_id,
                    affiliate_id=affiliate_id,
                    inv_date=inv_date,
                    due_date=due_date,
                    amount=amount,
                )
                inv_id = self.conn.execute(
                    "SELECT id FROM invoices WHERE number = ?", (invoice_number,)
                ).fetchone()["id"]

                # Create Mediciel payment if reglement > 0
                if reglement > 0:
                    self.conn.execute(
                        "INSERT INTO payments (invoice_id, client_id, date, amount, mode, reference) VALUES (?, ?, ?, ?, ?, ?)",
                        (inv_id, client_id, inv_date.isoformat(), reglement, "VIREMENT", f"MEDICIEL-{invoice_number}"),
                    )
                    self.conn.commit()
                    self.invoice_repo.update_status_from_payments(inv_id)

                stats["imported"] += 1

            except Exception:
                stats["errors"] += 1

        return stats
